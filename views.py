import datetime
import time
import ldap
import logging
import json
import re

from flask import render_template, request, redirect, url_for, session, abort, send_file

from flask.views import MethodView
import flask_login
from flask_login import login_required

from settings import API_SERVER

import utils
import forms
from models import User, TimeRecord


class UserAwareView(MethodView):
    """
    A base view class to extend.
    """

    @property
    def session(self):
        """
        Adds the session property to the view.
        """
        return session

    @property
    def user(self):
        """
        Adds the user property to the view.

        :returns: The currently logged in user if one exists, else None
        """
        if not flask_login.current_user.is_anonymous():
            return flask_login.current_user._get_current_object()
        else:
            return None

    def get_context(self, extra_ctx=None, **kwargs):
        """
        Adds a helper function to the view to get the context.

        :returns: The current context with the user set.
        """
        ctx = {
            'user': self.user,
        }
        if extra_ctx:
            ctx.update(extra_ctx)
        ctx.update(kwargs)
        return ctx


class Home(UserAwareView):
    """
    The view for the home page.
    """
    def get(self):
        context = self.get_context()
        context['nav'] = 'home'
        return render_template('index.html', **context)


class About(UserAwareView):
    """
    The view for the about page.
    """
    def get(self):
        context = self.get_context()
        context['nav'] = 'about'
        return render_template('about.html', **context)


class Login(UserAwareView):
    """
    The view for the login page.
    """
    def get(self):
        context = self.get_context()
        context['nav'] = 'login'
        context['form'] = forms.LoginForm()
        return render_template('login.html', **context)

    def post(self):
        form = forms.LoginForm(request.form)
        user = None

        username = form.username.data
        username = re.escape(username)
        password = form.password.data
        remember = form.remember_me.data

        if form.validate():
            try:
                logging.warning("Starting LDAP")
                conn = ldap.initialize('ldap://ldapaddress-here')
                conn.protocol_version = 3
                conn.set_option(ldap.OPT_REFERRALS, 0)
                conn.simple_bind_s(username + '@site-here.cdc.com', password)
            
                result_id = conn.search('DC=DC-here', ldap.SCOPE_SUBTREE, "(sAMAccountName=" + username + ")")
                                
                result_set = []
                while 1:
                    result_type, result_data = conn.result(result_id, 0)
                    if (result_data == []):
                        break
                    else:
                        if result_type == ldap.RES_SEARCH_ENTRY:
                            result_set.append(result_data)
                
                # At this point, we have gotten the user from
                # the AD server, and verified that they are
                # active
                if isActive or isTeamAdmin:
                    user = User.get_user_by_username(username)
                
                    if not user:
                        logging.debug("Creating User for: %s" % username)
                        user = User(username=username,
                                    is_approver=False,
                                    is_admin=False,
                                    ssn="00000000")
            
            except ldap.INVALID_CREDENTIALS:
                logging.warning("Invalid Credentials")
                user = None
            except ldap.SERVER_DOWN:
                logging.warning("Server down...")
                user = None
            
            if user:
                user.save()

                logging.debug("Authorized!")
                flask_login.login_user(user, remember=remember)
                return "success"
            
        return "Incorrect username or password"


class Logout(UserAwareView):
    """
    The view for the logout page.
    """
    decorators = [login_required]

    def get(self):
        flask_login.logout_user()
        return redirect(url_for('login'))


class Payroll(UserAwareView):
    """
    The view for the payroll page.
    """
    decorators = [login_required]
    
    def get(self, payroll_user=None, week=None):
        if payroll_user and not payroll_user == self.user.username:
            return redirect(url_for("payroll"))
        
        start_date = utils.get_last_monday(datetime.date.today())
        end_date = start_date + datetime.timedelta(days=6)
        if week:
            start_date = utils.get_last_monday(datetime.date.fromtimestamp(float(week)))
            end_date = start_date + datetime.timedelta(days=6)
            records = TimeRecord.get_current_week(self.user.username, start_date)
        else:
            records = TimeRecord.get_current_week(self.user.username)
        if not records:
            return abort(404)

        next_date = start_date + datetime.timedelta(days=7)
        prev_date = start_date - datetime.timedelta(days=7)
        context = {
            'nav':  'payroll',
            'user': self.user,
            'table_rows': records,
            'username': self.user.username,
            'start_date': start_date,
            'end_date': end_date,
            'prev_timestamp': time.mktime(prev_date.timetuple()),
            'next_timestamp': time.mktime(next_date.timetuple()),
        }
        return render_template('payroll.html', **context)

    def post(self, payroll_user=None, week=None):
        if payroll_user and not payroll_user == self.user.username:
            return redirect(url_for('payroll'))
        
        for input, value in request.form.iteritems():
            if value:
                punch_type, input_id = input.split('-')
                current_record = TimeRecord.objects(id=input_id).get()

                try:
                    time = datetime.datetime.strptime(value, '%I:%M %p')
                    day = current_record.date
                    timestamp = datetime.datetime.combine(day, time.time())
                except ValueError, e:
                    pass

                if punch_type == 'clockin':
                    current_record.clock_in = timestamp
                else:
                    current_record.clock_out = timestamp

                if current_record.clock_in and current_record.clock_out:
                    current_record.set_hours()

                current_record.save()
        if payroll_user and week:
            return redirect((url_for('payroll',
                                     payroll_user=payroll_user,
                                     week=week)))

        return redirect(url_for('payroll'))


class Approve(UserAwareView):
    """
    The view for the approve page.
    """
    decorators = [login_required]
    
    def get(self):
        if not self.user.is_approver:
            return redirect(url_for('payroll'))

        context = {
            'nav': 'approve',
            'user': self.user
        }

        records = TimeRecord.get_unapproved_records()

        context['records'] = records

        return render_template('approve.html', **context)

    def post(self):
        if not self.user.is_approver:
            return ""

        id = None
        approver = None
        if 'id' in request.form:
            approve, id = request.form['id'].split('-')
        if 'approver' in request.form:
            approver = request.form['approver']
        if not id or not approver:
            return "error"

        time_record = TimeRecord.objects(id=id).get()
        time_record.approved = True
        time_record.approved_by = self.user.username
        time_record.save()

        return approver


class Admin(UserAwareView):
    """
    The view for the admin page.
    """
    decorators = [login_required]
    
    def get(self):
        if not self.user.is_admin:
            return redirect(url_for('payroll'))
        
        users = User.objects(is_team_admin=False)
        add_user_form = forms.AddUser()
        context = {
            'nav': 'admin',
            'users': users,
            'user': self.user,
            'form': add_user_form,
            'api_server': API_SERVER
        }
        return render_template('admin.html', **context)

    def post(self):
        if not self.user.is_admin:
            return ""
        
        if not 'username' in request.form:
            return 'error'

        user = User.get_user_by_username(request.form['username'])
        for key, value in request.form.items():
            if hasattr(user, key):
                if value == 'true':
                    value = True
                elif value == 'false':
                    value = False
                setattr(user, key, value)
            user.save()

        data = {
            'user': user,
            'api_server': API_SERVER
        }

        return render_template('admin_user_row.html', **data)


class AddUser(UserAwareView):
    """
    The AJAX endpoint for adding a user to the system.
    """
    decorators = [login_required]
    
    def post(self):
        if not self.user.is_admin:
            return ""
        
        form = forms.AddUser(request.form)
        if form.validate():
            username = form.username.data
            is_admin = form.is_admin.data
            is_approver = form.is_approver.data
            ssn = form.ssn.data
            wage = form.wage.data

            user = User(username=username,
                        is_admin=is_admin,
                        is_approver=is_approver,
                        ssn=ssn,
                        wage=wage).save()

            data = {'user': user}

            return render_template('admin_user_row.html', **data)
        return 'error'


class DeleteUser(UserAwareView):
    """
    The AJAX endpoint for deleting a user from the system.
    """
    decorators = [login_required]
    
    def post(self):
        if not self.user.is_admin and not self.user.is_team_admin:
            return "error"

        username = None
        operator = None
        if 'username' in request.form:
            username = request.form['username']
        if 'operator' in request.form:
            operator = request.form['operator']
        if not username or not operator:
            return 'error'

        deleted = User.delete_user(str(username))
        return 'done'


class Export(UserAwareView):
    """
    The REST API endpoint for getting payroll info about a user.
    """
    
    def get(self, username):
        if not self.user or not self.user.is_admin:
            return redirect('http://www.site2.cdc.com/login')
        
        import tempfile
        from openpyxl import Workbook        
        days = int(request.args.get('days', 14))

        user = User.get_user_by_username(username)
        if not user:
            abort(404)

        records = TimeRecord.get_approved_records_by_username(user.username, num_days=days)

        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "Payroll Information"
    
        # User name
        ws.cell('%s%s' % ('A', 1)).value = 'User'
        ws.cell('%s%s' % ('A', 1)).style.font.bold = True
        ws.cell('%s%s' % ('B', 1)).value = user.username

        # SSN
        ws.cell('%s%s' % ('A', 2)).value = 'SSN'
        ws.cell('%s%s' % ('A', 2)).style.font.bold = True
        ws.cell('%s%s' % ('B', 2)).value = user.ssn
    
        # TimeRecord headers
        ws.cell('%s%s' % ('A', 4)).value = 'Date'
        ws.cell('%s%s' % ('A', 4)).style.font.bold = True
        ws.cell('%s%s' % ('B', 4)).value = 'Clock In'
        ws.cell('%s%s' % ('B', 4)).style.font.bold = True
        ws.cell('%s%s' % ('C', 4)).value = 'Clock Out'
        ws.cell('%s%s' % ('C', 4)).style.font.bold = True
        ws.cell('%s%s' % ('D', 4)).value = 'Approved?'
        ws.cell('%s%s' % ('D', 4)).style.font.bold = True
        ws.cell('%s%s' % ('E', 4)).value = 'Approved By'
        ws.cell('%s%s' % ('E', 4)).style.font.bold = True

        # All TimeRecords
        row = 5
        for record in records:
            ws.cell('%s%s' % ('A', row)).value = record.date.strftime('%B %d')
            ws.cell('%s%s' % ('B', row)).value = record.clock_in.strftime('%I:%M %p')
            ws.cell('%s%s' % ('C', row)).value = record.clock_out.strftime('%I:%M %p')
            ws.cell('%s%s' % ('D', row)).value = record.approved
            ws.cell('%s%s' % ('E', row)).value = record.approved_by
        
            row += 1

        filename = tempfile.mktemp()
        wb.save(filename=filename)

        return send_file(filename, 
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         attachment_filename="%s.xlsx" % user.username)
